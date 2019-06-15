import sys

import openpyxl
import requests
from bs4 import BeautifulSoup

import log


def escapeTUR(string):
    '''
    Escapes Turkish literals
    '''
    out = string[:]
    if "ö" in string:
        out = out.replace('ğ', "o")
    if "ü" in string:
        out = out.replace('ü', "u")
    if "ş" in string:
        out = out.replace('ş', "s")
    if 'ı' in string:
        out = out.replace('ı', "i")
    if 'ö' in string:
        out = out.replace('ö', "o")
    if 'ç' in string:
        out = out.replace('ç', "c")
    return out


def getSpecs(url: str, logger=None, verbose=False):
    '''
    Scrapes the spec sheet of the notebook in key, value pairs
    '''
    if logger:
        logger = log.Logger(log.DEBUG)

    # Process given URL
    try:
        with requests.get(url) as response:
            soup = BeautifulSoup(response.text, "html.parser")
            price = soup.find(
                "span", {"class": "urunDetay_satisFiyat"}).text.split(" ")[0]
            spec_cells = soup.find("div", {"class": "urunOzellik"}).findAll(
                "td", {"class": "gridAlternateDefault gridAlternateUrunOzellik"})
            manufacturer = soup.find(
                "h1", {"class": "emos_H1"}).text.split(" ")[0]
            name = soup.find("div", {"id": "plhUrunKodu"}).text.strip()
            content = []
            for cell in spec_cells:
                properties = {}
                properties["name"] = cell.find("div").text.strip()
                for pair in cell.findAll("tr"):
                    k = escapeTUR(pair.findAll("td")[
                                  0].text[:-6].strip().lower())
                    v = escapeTUR(pair.findAll("td")[
                                  1].text[:-6].strip().lower())
                    properties[k] = v
                content.append(properties)
    except ConnectionError:
        logger.critical("Failed to locate URL, ConnectionError has occurred.")
        logger.critical("Quitting...")
        return
    except Exception as ex:
        logger.critical("Unexpected type of error has occurred.")
        logger.trace(ex)
        logger.critical("Quitting...")
        return

    # If verbose mode is turned return all the properties
    if verbose:
        return content

    # Return desired properties as a tuple
    try:
        CPU = content[0]["i̇slemci teknolojisi"] + \
            "-" + content[0]["i̇slemci numarasi"]
        size = content[1]["ekran boyu"]
        resolution = content[1]["cozunurluk (piksel)"]
        storage = str(content[2]["disk kapasitesi"]) + \
            " " + str(content[2]["disk turu"])
        os = content[0]["i̇sletim sistemi"]
        return (manufacturer, name, price, CPU, size, resolution, storage, os)
    except KeyError:
        logger.error("Failed to locate properties, returning None.")
        return None
    except Exception as ex:
        logger.critical("Unexpected type of error has occurred.")
        logger.trace(ex)
        logger.critical("Quitting...")
        sys.exit(1)


def main():
    # Constants
    BASE_URL = "http://www.vatanbilgisayar.com"
    OUT_FILE = "outfile.xlsx"

    # Instantiate Logger Object and configure it
    logger = log.Logger(log.INFO)

    # XSLX Sheet
    try:
        workbook = openpyxl.load_workbook('template.xlsx')
        out_sheet = workbook[workbook.sheetnames[0]]
        logger.info("Template .xlsx file loaded successfully.")
    except FileNotFoundError as ex:
        logger.critical("Template .xlsx file not found.")
        logger.critical("Quitting...")
        return
    except Exception as ex:
        logger.critical("Unexpected type of error has occurred.")
        logger.trace(ex)
        logger.critical("Quitting...")
        sys.exit(1)

    # Grab product links
    links = []
    try:
        for i in range(1, 9):
            web_url = f"{BASE_URL}/notebook/?page={i}"
            main_page = requests.get(web_url)
            vatan = BeautifulSoup(main_page.text, "html.parser")
            logger.debug(f"Product list {i} returned {main_page.status_code}.")
            notebooks = vatan.findAll("div", {"class": "ems-prd-inner"})
            for notebook in notebooks[:-1]:
                url = notebook.find('div').a["href"]
                links.append(url)
            logger.info(f"Processed product list {i}.")
            logger.info(f"Total number of products is {len(links)}.")
    except ConnectionError:
        logger.critical("Failed to locate URL, ConnectionError has occurred.")
        logger.critical("Quitting...")
        return
    except Exception as ex:
        logger.critical("Unexpected type of error has occurred.")
        logger.trace(ex)
        logger.critical("Quitting...")
        sys.exit(1)
    logger.info(f"Processed all product lists total number of products is {len(links)}.")

    # Process product links
    row = 2
    error = {"count": 0, "links": []}
    for i, link in enumerate(links):
        # Gather data from the products page
        try:
            spec_url = f"{BASE_URL}{link}#urun-ozellikleri"
            result = getSpecs(spec_url, logger=logger)
            if result is None:
                error["count"] += 1
                error["links"].append(link)
                continue
            manufacturer, name, price, CPU, size, resolution, storage, os = result
            logger.info(f"Link number {i} specs processed.")
        except ConnectionError:
            logger.critical(
                "Failed to locate URL, ConnectionError has occurred.")
            logger.critical("Quitting...")
            return
        except Exception as ex:
            logger.critical("Unexpected type of error has occurred.")
            logger.trace(ex)
            logger.critical("Quitting...")
            sys.exit(1)

        # Write gathered data into the spreadsheet
        try:
            out_sheet.cell(row=row, column=1, value=manufacturer)
            out_sheet.cell(row=row, column=2, value=name)
            out_sheet.cell(row=row, column=3, value=price)
            out_sheet.cell(row=row, column=4, value=CPU)
            out_sheet.cell(row=row, column=5, value=size)
            out_sheet.cell(row=row, column=6, value=resolution)
            out_sheet.cell(row=row, column=7, value=storage)
            out_sheet.cell(row=row, column=8, value=os)
            out_sheet.cell(row=row, column=9, value=link)
            logger.info(f"Link number {i} written to the file.")
            row += 1
            workbook.save(OUT_FILE)
        except PermissionError:
            logger.critical(f"Permission denied when accessing file ({OUT_FILE}).")
            logger.critical("Quitting...")
            return
        except Exception as ex:
            logger.critical("Unexpected type of error has occurred.")
            logger.trace(ex)
            logger.critical("Quitting...")
            sys.exit(1)
    logger.debug(f'Skipped {error["count"]} links.')
    for link in error["links"]:
        logger.info(f"{BASE_URL}{link}#urun-ozellikleri")


if __name__ == "__main__":
    main()
